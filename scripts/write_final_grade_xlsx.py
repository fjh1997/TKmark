#!/usr/bin/env python3

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "排名",
    "学号",
    "姓名",
    "平时成绩",
    "平时原始分",
    "期末成绩",
    "最终成绩",
    "是否及格",
    "作业均分",
    "讨论加分",
    "考勤分",
    "未签到",
    "请假",
    "公假",
    "私假",
    "病假",
    "未分类请假",
    "迟到",
    "早退",
    "考勤备注",
    "请假明细",
]


def value(row: dict, key: str):
    return row.get(key, "")


def build_row(index: int, row: dict) -> list:
    return [
        index,
        str(value(row, "studentNo")),
        value(row, "name"),
        value(row, "ordinaryScore"),
        value(row, "ordinaryRawScore"),
        value(row, "examScore"),
        value(row, "finalScore"),
        "及格" if value(row, "passed") else "不及格",
        value(row, "assignmentAverage"),
        value(row, "discussionBonus"),
        value(row, "attendanceScore"),
        value(row, "notSignCount"),
        value(row, "leaveCount"),
        value(row, "publicLeaveCount"),
        value(row, "privateLeaveCount"),
        value(row, "sickLeaveCount"),
        value(row, "unknownLeaveCount"),
        value(row, "lateCount"),
        value(row, "earlyCount"),
        value(row, "attendanceNote"),
        value(row, "leaveReasons"),
    ]


def autosize(ws):
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 48)
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["U"].width = 48


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    fail_fill = PatternFill("solid", fgColor="FFF1F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        row[1].number_format = "@"
        row[1].value = str(row[1].value or "")
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column >= 20)
        if row[7].value == "不及格":
            for cell in row:
                cell.fill = fail_fill


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: write_final_grade_xlsx.py report.json report.xlsx", file=sys.stderr)
        return 2

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    data = json.loads(input_path.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)
    for class_name, rows in data["classes"].items():
        ws = wb.create_sheet(title=class_name[:31])
        ws.append(HEADERS)
        for index, row in enumerate(rows, start=1):
            ws.append(build_row(index, row))
        style_sheet(ws)
        autosize(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
