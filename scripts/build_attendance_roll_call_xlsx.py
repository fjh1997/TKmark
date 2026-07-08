#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


DEFAULT_CLASS_TEMPLATES = {
    "信安24-01": "2401学生成绩录入模板[傅继晗].xlsx",
    "信安24实验班": "实验班学生成绩录入模板[傅继晗].xlsx",
    "信安2504": "2504学生成绩录入模板[傅继晗].xlsx",
}

DEFAULT_MARKS = {
    "present": "/",
    "absent": "O",
    "sick_leave": "⊙",
    "private_leave": "⊕",
    "late": "◎",
    "early": "X",
    "public_leave": "公",
    "unknown_leave": "假",
}

ROWS_PER_PAGE = 25
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build printable attendance roll-call xlsx.")
    parser.add_argument(
        "--report-json",
        default="out/final-grades-20260706/final-grade-report.json",
        help="final-grade-report.json path",
    )
    parser.add_argument(
        "--template-dir",
        default="../期末成绩录入模板",
        help="directory containing class score-entry templates",
    )
    parser.add_argument("--course-name", default="信息安全代码审计", help="course name shown in the form")
    parser.add_argument("--out", default="", help="output xlsx path")
    return parser.parse_args()


def read_report(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def roster_from_template(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["学生成绩录入模板"]
    roster = []
    for row_index in range(2, ws.max_row + 1):
        student_no = str(ws.cell(row_index, 3).value or "").strip()
        name = str(ws.cell(row_index, 4).value or "").strip()
        if student_no:
            roster.append({"studentNo": student_no, "name": name})
    return roster


def parse_leave_reasons(text: str) -> list[tuple[str, str]]:
    items = []
    for part in re.split(r"[；;]", str(text or "")):
        part = part.strip()
        if not part or ":" not in part:
            continue
        date, reason = part.split(":", 1)
        date = date.strip()
        reason = reason.strip() or "未填写原因"
        if re.fullmatch(r"20\d{2}-\d{2}-\d{2}", date):
            items.append((date, reason))
    return items


def mark_for_leave(reason: str, marks: dict) -> str:
    if re.search(r"公|因公|公务|比赛|竞赛|学校|学院", reason):
        return marks["public_leave"]
    if "病" in reason:
        return marks["sick_leave"]
    if re.search(r"私|事假", reason):
        return marks["private_leave"]
    return marks["unknown_leave"]


def mark_for_attendance_event(event: dict, marks: dict) -> str:
    status = str(event.get("status") or "").strip()
    reason = str(event.get("leaveReason") or "").strip()
    status_text = status + " " + str(event.get("statusName") or event.get("signStatus") or "")

    if status == "2" or "请假" in status_text:
        return mark_for_leave(reason, marks)
    if status in {"3", "late"} or "迟到" in status_text:
        return marks["late"]
    if status in {"4", "early"} or "早退" in status_text:
        return marks["early"]
    if status in {"0", "absent"} or "未签到" in status_text or "缺勤" in status_text:
        return marks["absent"]
    return marks["present"]


def build_class_attendance(class_name: str, roster: list[dict], rows_by_no: dict, marks: dict):
    dates = set()
    marks_by_no: dict[str, dict[str, str]] = {}

    for student in roster:
        student_no = student["studentNo"]
        row = rows_by_no.get(student_no, {})
        date_marks = {}

        events = row.get("attendanceEvents") or []
        for event in events:
            date = str(event.get("date") or "").strip()
            if not re.fullmatch(r"20\d{2}-\d{2}-\d{2}", date):
                continue
            dates.add(date)
            date_marks[date] = mark_for_attendance_event(event, marks)

        if not events:
            for date, reason in parse_leave_reasons(row.get("leaveReasons", "")):
                dates.add(date)
                date_marks[date] = mark_for_leave(reason, marks)

        marks_by_no[student_no] = date_marks

    return sorted(dates), marks_by_no


def attendance_note(row: dict) -> str:
    parts = []
    leave_note = str(row.get("attendanceNote") or "").strip()
    if leave_note:
        parts.append(leave_note)

    not_sign = int(row.get("notSignCount") or 0)
    late = int(row.get("lateCount") or 0)
    early = int(row.get("earlyCount") or 0)
    if not_sign:
        parts.append(f"缺课{not_sign}（无日期明细）")
    if late:
        parts.append(f"迟到{late}（无日期明细）")
    if early:
        parts.append(f"早退{early}（无日期明细）")
    return "；".join(parts)


def border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = CELL_BORDER


def write_page(
    ws,
    class_name: str,
    course_name: str,
    roster_slice: list[dict],
    rows_by_no: dict,
    page_index: int,
    start_row: int,
    dates: list[str],
    marks_by_no: dict,
    marks: dict,
) -> tuple[int, int]:
    date_count = max(1, len(dates))
    first_date_col = 4
    last_date_col = first_date_col + date_count - 1
    remark_col = last_date_col + 1
    total_cols = remark_col

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    title_cell = ws.cell(start_row, 1, "学生点名册")
    title_cell.font = Font(name="宋体", size=18, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 34

    info_row = start_row + 1
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=3)
    ws.cell(info_row, 1, "（学生）系：")
    ws.merge_cells(start_row=info_row, start_column=4, end_row=info_row, end_column=min(7, total_cols))
    ws.cell(info_row, 4, f"班级：{class_name}")
    if total_cols >= 8:
        ws.merge_cells(start_row=info_row, start_column=8, end_row=info_row, end_column=total_cols)
        ws.cell(info_row, 8, f"课程名称：{course_name}")
    for col in range(1, total_cols + 1):
        ws.cell(info_row, col).font = Font(name="宋体", size=12)
        ws.cell(info_row, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[info_row].height = 26

    header_row = start_row + 2
    date_row = start_row + 3
    data_start = start_row + 4
    data_end = data_start + ROWS_PER_PAGE - 1
    note_row = data_end + 1

    ws.merge_cells(start_row=header_row, start_column=1, end_row=date_row, end_column=1)
    ws.merge_cells(start_row=header_row, start_column=2, end_row=date_row, end_column=2)
    ws.merge_cells(start_row=header_row, start_column=3, end_row=date_row, end_column=3)
    ws.cell(header_row, 1, "序\n号")
    ws.cell(header_row, 2, "姓名")
    ws.cell(header_row, 3, "性\n别")
    ws.merge_cells(start_row=header_row, start_column=first_date_col, end_row=header_row, end_column=last_date_col)
    ws.cell(header_row, first_date_col, "上  课  日  期")
    ws.merge_cells(start_row=header_row, start_column=remark_col, end_row=date_row, end_column=remark_col)
    ws.cell(header_row, remark_col, "备注")

    for index, date in enumerate(dates or [""]):
        cell = ws.cell(date_row, first_date_col + index, date[5:] if date else "")
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        cell.font = Font(name="宋体", size=8)

    for col in range(1, total_cols + 1):
        ws.cell(header_row, col).fill = HEADER_FILL
        ws.cell(date_row, col).fill = HEADER_FILL
        ws.cell(header_row, col).font = Font(name="宋体", size=11, bold=True)
        ws.cell(date_row, col).font = Font(name="宋体", size=8, bold=True)
        ws.cell(header_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col not in range(first_date_col, last_date_col + 1):
            ws.cell(date_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index in range(ROWS_PER_PAGE):
        row_index = data_start + index
        sequence = page_index * ROWS_PER_PAGE + index + 1
        if index < len(roster_slice):
            student = roster_slice[index]
            student_no = student["studentNo"]
            source_row = rows_by_no.get(student_no, {})
            ws.cell(row_index, 1, sequence)
            ws.cell(row_index, 2, student["name"])
            ws.cell(row_index, 3, "")
            student_marks = marks_by_no.get(student_no, {})
            for date_index, date in enumerate(dates or [""]):
                value = student_marks.get(date, marks["present"] if date else "")
                ws.cell(row_index, first_date_col + date_index, value)
            ws.cell(row_index, remark_col, attendance_note(source_row))
        for col in range(1, total_cols + 1):
            ws.cell(row_index, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row_index, col).font = Font(name="宋体", size=10)
        ws.cell(row_index, remark_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_index].height = 24

    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=total_cols)
    note = (
        f"注：到课{marks['present']}  缺课{marks['absent']}  病假{marks['sick_leave']}  "
        f"事假{marks['private_leave']}  公假{marks['public_leave']}  "
        f"未填原因请假{marks['unknown_leave']}  迟到{marks['late']}  早退{marks['early']}"
    )
    ws.cell(note_row, 1, note)
    ws.cell(note_row, 1).font = Font(name="宋体", size=11, bold=True)
    ws.cell(note_row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[note_row].height = 24

    border_range(ws, header_row, data_end, 1, total_cols)
    for col in range(1, total_cols + 1):
        ws.cell(header_row, col).border = Border(left=THIN, right=THIN, top=MEDIUM, bottom=THIN)
        ws.cell(data_end, col).border = Border(left=THIN, right=THIN, top=THIN, bottom=MEDIUM)
    for row in range(header_row, data_end + 1):
        ws.cell(row, 1).border = Border(left=MEDIUM, right=THIN, top=ws.cell(row, 1).border.top, bottom=ws.cell(row, 1).border.bottom)
        ws.cell(row, total_cols).border = Border(
            left=THIN,
            right=MEDIUM,
            top=ws.cell(row, total_cols).border.top,
            bottom=ws.cell(row, total_cols).border.bottom,
        )

    return note_row + 2, total_cols


def build_workbook(report: dict, template_dir: Path, course_name: str) -> Workbook:
    rows_by_class = {
        class_name: {str(row["studentNo"]).strip(): row for row in rows}
        for class_name, rows in report["classes"].items()
    }

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("说明")
    summary.append(["项目", "说明"])
    summary.append(["数据来源", "final-grade-report.json"])
    summary.append(["生成说明", "优先使用 attendanceEvents 逐次考勤；如果没有逐次事件，则使用 leaveReasons 中的请假日期。"])
    summary.append(["限制", "只有汇总次数且没有具体日期的缺课、迟到、早退会写入备注列。"])
    summary.append(
        [
            "标记",
            f"到课{DEFAULT_MARKS['present']} 缺课{DEFAULT_MARKS['absent']} "
            f"病假{DEFAULT_MARKS['sick_leave']} 事假{DEFAULT_MARKS['private_leave']} "
            f"公假{DEFAULT_MARKS['public_leave']} 未填原因请假{DEFAULT_MARKS['unknown_leave']} "
            f"迟到{DEFAULT_MARKS['late']} 早退{DEFAULT_MARKS['early']}",
        ]
    )
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 100

    for class_name, template_name in DEFAULT_CLASS_TEMPLATES.items():
        template_path = template_dir / template_name
        if not template_path.exists():
            raise FileNotFoundError(f"template not found: {template_path}")

        roster = roster_from_template(template_path)
        class_rows = rows_by_class.get(class_name, {})
        dates, marks_by_no = build_class_attendance(class_name, roster, class_rows, DEFAULT_MARKS)
        ws = workbook.create_sheet(class_name[:31])

        next_row = 1
        max_cols = 1
        page_count = (len(roster) + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE
        for page_index in range(page_count):
            roster_slice = roster[page_index * ROWS_PER_PAGE : (page_index + 1) * ROWS_PER_PAGE]
            next_row, cols = write_page(
                ws,
                class_name,
                course_name,
                roster_slice,
                class_rows,
                page_index,
                next_row,
                dates,
                marks_by_no,
                DEFAULT_MARKS,
            )
            max_cols = max(max_cols, cols)
            if page_index < page_count - 1:
                ws.row_breaks.append(Break(id=next_row - 1))

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 5
        for col in range(4, max_cols):
            ws.column_dimensions[get_column_letter(col)].width = 4.2
        ws.column_dimensions[get_column_letter(max_cols)].width = 24
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.print_options.horizontalCentered = True

    return workbook


def main() -> int:
    args = parse_args()
    report_path = Path(args.report_json)
    template_dir = Path(args.template_dir)
    output_path = Path(args.out) if args.out else Path("out/attendance-roll-call") / f"三班考勤点名册_{datetime.now():%Y%m%d-%H%M%S}.xlsx"

    report = read_report(report_path)
    workbook = build_workbook(report, template_dir, args.course_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
